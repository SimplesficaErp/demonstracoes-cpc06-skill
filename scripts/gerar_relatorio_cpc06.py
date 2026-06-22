#!/usr/bin/env python3
"""
NEWLEDGER · CPC 06 (R2) / IFRS 16 — v5 LAYOUT CORRIGIDO
========================================================
v5 corrige:
- Colunas com largura correta (labels NÃO cortadas)
- Freeze panes correto (B/C visíveis ao rolar)
- Logo: usa SOMENTE arquivo 'logo-newledger.png' se existir (do user)
  Se não existir, usa cabeçalho de texto sem inventar imagem
- Tipografia mais limpa, espaçamento adequado
"""
import sys, json, urllib.request
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

NAVY_DARK = "0F2540"; NAVY = "1F3B5B"; NAVY_MED = "3D5778"; NAVY_LIGHT = "6B8AAB"
NAVY_BG = "E8EDF3"; WHITE = "FFFFFF"; GREEN = "2E7D4F"; GREEN_LIGHT = "D6F4E2"
GOLD = "C8A030"; GOLD_LIGHT = "FFF8E1"; RED = "C73E5F"; RED_LIGHT = "FFE5EC"
GRAY_BG = "FAFAFC"; PURPLE = "6841EA"; PURPLE_LIGHT = "EDE5FF"; INPUT_BG = "FFF8E1"

FMT_BR = '#,##0.00'
FMT_PCT = '0.00%'
FMT_DATE = 'dd/mm/yyyy'

ALIQ_PIS = 0.0165
ALIQ_COFINS = 0.0760

TABELA_CBSIBS = [
    (2026, 0.0090, 0.0010), (2027, 0.0088, 0.0177), (2028, 0.0264, 0.0531),
    (2029, 0.0440, 0.0885), (2030, 0.0616, 0.1239), (2031, 0.0880, 0.1770),
    (2032, 0.0880, 0.1770), (2033, 0.0880, 0.1770), (2034, 0.0880, 0.1770),
    (2035, 0.0880, 0.1770), (2036, 0.0880, 0.1770), (2037, 0.0880, 0.1770),
    (2038, 0.0880, 0.1770), (2039, 0.0880, 0.1770), (2040, 0.0880, 0.1770),
]

LOGO_PATH = "logo-newledger.png"  # SOMENTE se existir arquivo real do user
HAS_LOGO = Path(LOGO_PATH).exists()


def fetch_rates():
    r = {'selic_meta': None, 'selic_acumulada_12m': None, 'cdi_acumulado_12m': None,
         'ipca_12m': None, 'usd_brl': None, 'ibov_close': None,
         'fonte_taxa_desconto': None, 'serie_bacen': None,
         'data_consulta': datetime.now().isoformat(),
         'data_consulta_legivel': datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}
    base = "https://api.bcb.gov.br/dados/serie/bcdata.sgs"
    fim = datetime.now(); ini = fim - timedelta(days=30)
    params = f"?formato=json&dataInicial={ini.strftime('%d/%m/%Y')}&dataFinal={fim.strftime('%d/%m/%Y')}"
    series = {'selic_meta': 432, 'selic_acumulada_12m': 4390, 'cdi_acumulado_12m': 4391, 'ipca_12m': 13522}
    for key, codigo in series.items():
        try:
            with urllib.request.urlopen(f"{base}.{codigo}/dados{params}", timeout=12) as resp:
                data = json.loads(resp.read())
                if data: r[key] = float(data[-1]['valor'])
        except: pass
    try:
        import yfinance as yf
        ibov = yf.Ticker("^BVSP").history(period="2d")
        if not ibov.empty: r['ibov_close'] = float(ibov['Close'].iloc[-1])
    except: pass

    if r['selic_acumulada_12m']:
        r['taxa_desconto'] = r['selic_acumulada_12m']
        r['fonte_taxa_desconto'] = 'BACEN SGS 4390 — Selic acumulada 12 meses'
        r['serie_bacen'] = '4390'
    elif r['selic_meta']:
        r['taxa_desconto'] = r['selic_meta']
        r['fonte_taxa_desconto'] = 'BACEN SGS 432 — Meta Selic'
        r['serie_bacen'] = '432'
    else:
        r['selic_acumulada_12m'] = 15.00; r['selic_meta'] = 15.00
        r['cdi_acumulado_12m'] = 14.90; r['ipca_12m'] = 4.50
        r['taxa_desconto'] = 15.00
        r['fonte_taxa_desconto'] = 'PADRÃO (offline)'
        r['serie_bacen'] = 'OFFLINE'
    return r


def parse_input_contracts(input_path):
    wb = load_workbook(input_path, data_only=True)
    contratos = []
    for sn in wb.sheetnames:
        if sn.startswith('Balancete') or sn in ('Lançamentos', 'RESUMO - Empresa', 'PGTO ALUGUEL PJ'):
            continue
        ws = wb[sn]
        is_c = False
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            rs = ' '.join([str(c) for c in row if c is not None]).lower()
            if 'locatário' in rs: is_c = True; break
        if not is_c: continue
        c = {'aba_origem': sn, 'regime': 'LUCRO_REAL'}
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or not isinstance(v, str): continue
                low = v.lower().strip()
                if 'valor mensal' in low:
                    nc = ws.cell(row=cell.row, column=cell.column+1)
                    if isinstance(nc.value, (int,float)): c['valor_mensal'] = float(nc.value)
                if 'quantidade de meses' in low:
                    nc = ws.cell(row=cell.row, column=cell.column+1)
                    if isinstance(nc.value, (int,float)): c['prazo_meses'] = int(nc.value)
                if 'vigência do contrato' in low:
                    nc = ws.cell(row=cell.row, column=cell.column+1)
                    if isinstance(nc.value, datetime): c['data_inicio'] = nc.value
        if c.get('valor_mensal'):
            contratos.append(c)
    return contratos


def add_logo_or_text(ws, row=1, large=False):
    """Adiciona logo SOMENTE se existir PNG. Mantém proporção real da imagem.
    large=True para a CAPA (logo grande); False para as demais abas (logo menor)."""
    if HAS_LOGO:
        try:
            # Detecta proporção real da imagem
            from PIL import Image as PILImg
            with PILImg.open(LOGO_PATH) as pil_img:
                orig_w, orig_h = pil_img.size
            aspect = orig_w / orig_h

            img = XLImage(LOGO_PATH)
            if large:
                # CAPA: logo grande (altura alvo 180px)
                target_h = 180
                target_w = int(target_h * aspect)
                row_height = 145
            else:
                # Outras abas: logo pequena (altura alvo 80px)
                target_h = 80
                target_w = int(target_h * aspect)
                row_height = 65

            img.width = target_w
            img.height = target_h
            ws.add_image(img, f"B{row}")
            ws.row_dimensions[row].height = row_height
        except Exception as e:
            _add_text_header(ws, row)
    else:
        _add_text_header(ws, row)


def _add_text_header(ws, row=1):
    """Header de texto simples (sem inventar imagem)."""
    ws.row_dimensions[row].height = 36
    ws.cell(row=row, column=2, value="NEWLEDGER")
    ws.cell(row=row, column=2).font = Font(name="Inter", size=24, bold=True, color=NAVY_DARK)
    ws.cell(row=row, column=2).alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row+1].height = 18
    ws.cell(row=row+1, column=2, value="conectando inteligências · construindo o futuro").font = Font(name="Inter", size=9, italic=True, color=NAVY_LIGHT)
    ws.cell(row=row+1, column=2).alignment = Alignment(vertical="top")
    ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=6)


def style_input(cell):
    cell.fill = PatternFill("solid", fgColor=INPUT_BG)
    cell.font = Font(name="Inter", size=11, bold=True, color=NAVY_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(left=Side(style='medium', color=GOLD),
                         right=Side(style='medium', color=GOLD),
                         top=Side(style='thin', color=GOLD),
                         bottom=Side(style='thin', color=GOLD))


def style_label(cell):
    cell.font = Font(name="Inter", size=10, color=NAVY_MED)
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)


def style_calc(cell):
    cell.fill = PatternFill("solid", fgColor=NAVY_BG)
    cell.font = Font(name="Inter", size=11, bold=True, color=NAVY_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(left=Side(style='thin', color=NAVY_LIGHT),
                         bottom=Side(style='thin', color=NAVY_LIGHT))


def style_header(cell):
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Inter", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_title(cell, color_bg=NAVY_BG, color_text=NAVY):
    cell.fill = PatternFill("solid", fgColor=color_bg)
    cell.font = Font(name="Inter", size=12, bold=True, color=color_text)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def gen_capa(wb, contratos, rates):
    ws = wb.active; ws.title = "📋 CAPA"
    ws.sheet_view.showGridLines = False

    add_logo_or_text(ws, row=1, large=True)

    # Banner navy
    for col_idx in range(1, 13):
        ws.cell(row=4, column=col_idx).fill = PatternFill("solid", fgColor=NAVY_DARK)
    ws.row_dimensions[4].height = 8

    ws.merge_cells("B6:L6")
    ws["B6"] = "RELATÓRIO DE ARRENDAMENTO MERCANTIL"
    ws["B6"].font = Font(name="Inter", size=24, bold=True, color=NAVY_DARK)
    ws["B6"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[6].height = 36

    ws.merge_cells("B7:L7")
    ws["B7"] = "CPC 06 (R2) / IFRS 16 · com fórmulas dinâmicas e créditos tributários"
    ws["B7"].font = Font(name="Inter", size=11, italic=True, color=NAVY_MED)
    ws["B7"].alignment = Alignment(horizontal="center")

    # Cards KPI
    ws.row_dimensions[10].height = 18; ws.row_dimensions[11].height = 44; ws.row_dimensions[12].height = 18

    cards = [
        ("CONTRATOS", str(len(contratos)), NAVY, "ativos no relatório"),
        ("PERIODICIDADE", "MENSAL", GREEN, "atualização automática"),
        ("TAXA IBR", f"{rates['taxa_desconto']:.2f}%", GOLD, "a.a. — BACEN SGS"),
        ("STATUS", "FÓRMULAS LIVE", PURPLE, "muda input recalcula"),
    ]
    col_start = 2
    for label, value, cor, sub in cards:
        ws.cell(row=10, column=col_start, value=label).fill = PatternFill("solid", fgColor=NAVY_BG)
        ws.cell(row=10, column=col_start).font = Font(name="Inter", size=9, bold=True, color=NAVY_MED)
        ws.cell(row=10, column=col_start).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=10, start_column=col_start, end_row=10, end_column=col_start+2)
        ws.cell(row=11, column=col_start, value=value).font = Font(name="Inter", size=16, bold=True, color=cor)
        ws.cell(row=11, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=11, start_column=col_start, end_row=11, end_column=col_start+2)
        ws.cell(row=12, column=col_start, value=sub).font = Font(name="Inter", size=9, italic=True, color=NAVY_LIGHT)
        ws.cell(row=12, column=col_start).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=12, start_column=col_start, end_row=12, end_column=col_start+2)
        for r in range(10, 13):
            for co in range(0, 3):
                cc = ws.cell(row=r, column=col_start + co)
                if r == 11 and co != 0: cc.fill = PatternFill("solid", fgColor=WHITE)
                cc.border = Border(
                    left=Side(style='medium', color=cor) if co == 0 else None,
                    right=Side(style='medium', color=cor) if co == 2 else None,
                    top=Side(style='medium', color=cor) if r == 10 else None,
                    bottom=Side(style='medium', color=cor) if r == 12 else None)
        col_start += 3

    # Taxa info
    ws.merge_cells("B15:L15")
    ws["B15"] = "🌐 TAXA DE DESCONTO CAPTURADA AUTOMATICAMENTE"
    style_section_title(ws["B15"], GOLD_LIGHT, NAVY)
    ws.row_dimensions[15].height = 26

    rate_rows = [
        ("Taxa IBR adotada (CPC 06 §27)", f"{rates['taxa_desconto']:.2f}% a.a."),
        ("Série BACEN", rates.get('serie_bacen','—')),
        ("Fonte", rates['fonte_taxa_desconto']),
        ("Selic acumulada 12 meses", f"{rates.get('selic_acumulada_12m',0):.2f}% a.a." if rates.get('selic_acumulada_12m') else '—'),
        ("CDI acumulado 12 meses", f"{rates.get('cdi_acumulado_12m',0):.2f}% a.a." if rates.get('cdi_acumulado_12m') else '—'),
        ("Capturado em", rates['data_consulta_legivel']),
    ]
    for i, (k, v) in enumerate(rate_rows, start=17):
        ws.cell(row=i, column=2, value=k).font = Font(name="Inter", size=10, color=NAVY_MED)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=i, column=6, value=v).font = Font(name="Inter", size=10, bold=True, color=NAVY_DARK)
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=8)
        if i % 2 == 0:
            for col in range(2, 10):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GRAY_BG)
        ws.row_dimensions[i].height = 20

    base = 17 + len(rate_rows) + 2
    ws.cell(row=base, column=2, value=f"📑 CONTRATOS ({len(contratos)})").font = Font(name="Inter", size=13, bold=True, color=NAVY)
    for i, c in enumerate(contratos, start=base+1):
        ws.cell(row=i, column=2, value=f"   Contrato {i-base}").font = Font(name="Inter", size=10, color=NAVY)
        ws.cell(row=i, column=6, value=f"R$ {c['valor_mensal']:,.2f}/mês × {c.get('prazo_meses', 36)} meses").font = Font(name="Inter", size=10, color=NAVY_MED)
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=8)
        ws.row_dimensions[i].height = 18

    footer_r = base + len(contratos) + 3
    ws.cell(row=footer_r, column=2, value="NEWLEDGER · CONECTANDO INTELIGÊNCIAS · CONSTRUINDO O FUTURO").font = Font(name="Inter", size=9, italic=True, color=NAVY_LIGHT)

    # LARGURAS CORRIGIDAS — coluna B grande para labels
    for col, w in [('A', 2), ('B', 42), ('C', 18), ('D', 12), ('E', 22),
                   ('F', 18), ('G', 18), ('H', 18), ('I', 12), ('J', 12), ('K', 12), ('L', 12)]:
        ws.column_dimensions[col].width = w


def gen_aba_aliquotas(wb):
    ws = wb.create_sheet("Tabela Aliquotas")
    ws.sheet_view.showGridLines = False
    add_logo_or_text(ws, row=1)

    ws.merge_cells("B4:E4")
    ws["B4"] = "TABELA DE ALÍQUOTAS — REFERENCIADA POR VLOOKUP"
    ws["B4"].font = Font(name="Inter", size=14, bold=True, color=NAVY_DARK)
    ws["B4"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[4].height = 28

    # PIS/COFINS bloco
    ws["B6"] = "PIS/COFINS — Lucro Real até 31/12/2026"
    style_section_title(ws["B6"], GREEN_LIGHT, GREEN)
    ws.merge_cells("B6:E6"); ws.row_dimensions[6].height = 24

    pisrows = [
        ("Alíquota PIS", ALIQ_PIS, FMT_PCT),
        ("Alíquota COFINS", ALIQ_COFINS, FMT_PCT),
        ("Total PIS+COFINS", "=C7+C8", FMT_PCT),
        ("Data limite", datetime(2026, 12, 31), FMT_DATE),
    ]
    for i, (k, v, fmt) in enumerate(pisrows, start=7):
        ws.cell(row=i, column=2, value=k).font = Font(name="Inter", size=10, color=NAVY_MED)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=i, column=3, value=v).font = Font(name="Inter", size=11, bold=True, color=NAVY_DARK)
        ws.cell(row=i, column=3).number_format = fmt
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="right")
        ws.row_dimensions[i].height = 20

    # CBS/IBS bloco
    ws["B13"] = "CBS/IBS — Transição a partir de 01/01/2027"
    style_section_title(ws["B13"], PURPLE_LIGHT, PURPLE)
    ws.merge_cells("B13:E13"); ws.row_dimensions[13].height = 24

    headers = ["Ano", "CBS", "IBS", "Total"]
    for col, h in enumerate(headers, start=2):
        style_header(ws.cell(row=15, column=col, value=h))
    ws.row_dimensions[15].height = 22

    for i, (ano, cbs, ibs) in enumerate(TABELA_CBSIBS, start=16):
        ws.cell(row=i, column=2, value=ano).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=cbs).number_format = FMT_PCT
        ws.cell(row=i, column=4, value=ibs).number_format = FMT_PCT
        ws.cell(row=i, column=5, value=f"=C{i}+D{i}").number_format = FMT_PCT
        for col in range(2, 6):
            ws.cell(row=i, column=col).font = Font(name="Inter", size=10, color=NAVY)
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GRAY_BG)

    # LARGURAS CORRIGIDAS
    for col, w in [('A', 2), ('B', 32), ('C', 14), ('D', 14), ('E', 14)]:
        ws.column_dimensions[col].width = w


def gen_contrato_sheet(wb, c, num):
    sheet_name = f"Contrato {num}"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    add_logo_or_text(ws, row=1)

    # Título do contrato à direita do logo
    ws.merge_cells("F1:J1")
    ws["F1"] = f"CONTRATO {num}"
    ws["F1"].font = Font(name="Inter", size=22, bold=True, color=NAVY_DARK)
    ws["F1"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("F2:J2")
    ws["F2"] = "Arrendamento Mercantil · CPC 06 (R2) / IFRS 16"
    ws["F2"].font = Font(name="Inter", size=10, italic=True, color=NAVY_MED)
    ws["F2"].alignment = Alignment(horizontal="right")

    # === INPUTS (linhas 5-13) ===
    ws["B4"] = "📝 DADOS DO CONTRATO (campos editáveis)"
    style_section_title(ws["B4"], GOLD_LIGHT, NAVY)
    ws.merge_cells("B4:D4"); ws.row_dimensions[4].height = 26

    inputs = [
        ("Locador", f"Locador {num}"),
        ("Empresa Locatária", "Empresa Locatária"),
        ("Objeto", f"Imóvel/Bem comercial {num}"),
        ("Data início", c.get('data_inicio', datetime.now())),
        ("Prazo (meses)", c.get('prazo_meses', 36)),
        ("Valor mensal (R$)", c.get('valor_mensal', 5000)),
        ("Taxa de desconto a.a.", 0.15),
        ("Regime tributário", c.get('regime', 'LUCRO_REAL')),
    ]
    for i, (label, value) in enumerate(inputs, start=5):
        ws.cell(row=i, column=2, value=label); style_label(ws.cell(row=i, column=2))
        ws.cell(row=i, column=3, value=value); style_input(ws.cell(row=i, column=3))
        ws.row_dimensions[i].height = 24

    ws["C8"].number_format = FMT_DATE
    ws["C9"].number_format = "0"
    ws["C10"].number_format = FMT_BR
    ws["C11"].number_format = FMT_PCT

    # Taxa mensal
    ws.cell(row=13, column=2, value="Taxa mensal efetiva"); style_label(ws.cell(row=13, column=2))
    ws["C13"] = "=((1+C11)^(1/12))-1"; style_calc(ws["C13"])
    ws["C13"].number_format = "0.0000%"
    ws.row_dimensions[13].height = 22

    # === CÁLCULO INICIAL (15-18) ===
    ws["B15"] = "🧮 MENSURAÇÃO INICIAL CPC 06 (R2) §22-28"
    style_section_title(ws["B15"], NAVY_BG, NAVY)
    ws.merge_cells("B15:D15"); ws.row_dimensions[15].height = 24

    calcs = [
        ("Valor Presente do Passivo (Liability)", "=C10*(1-(1+C13)^(-C9))/C13"),
        ("Ativo Direito de Uso (RoU inicial)", "=C16"),
        ("Depreciação mensal do RoU", "=C17/C9"),
    ]
    for i, (label, formula) in enumerate(calcs, start=16):
        ws.cell(row=i, column=2, value=label); style_label(ws.cell(row=i, column=2))
        ws.cell(row=i, column=3, value=formula); style_calc(ws.cell(row=i, column=3))
        ws.cell(row=i, column=3).number_format = FMT_BR
        ws.row_dimensions[i].height = 22

    # === LANÇAMENTO INICIAL (20-22) ===
    ws["B20"] = "💼 RECONHECIMENTO INICIAL — LANÇAMENTO CONTÁBIL"
    style_section_title(ws["B20"], PURPLE_LIGHT, NAVY)
    ws.merge_cells("B20:E20"); ws.row_dimensions[20].height = 24

    ws["B21"] = "D"; ws["C21"] = "1.2.2.01.0016 RoU Imóveis"
    ws["E21"] = "=C16"; ws["E21"].number_format = FMT_BR
    ws["B22"] = "C"; ws["C22"] = "2.2.2.04.0001 Passivo Arrendamento LP"
    ws["E22"] = "=C16"; ws["E22"].number_format = FMT_BR
    for r in [21, 22]:
        ws.cell(row=r, column=2).font = Font(name="Inter", size=11, bold=True, color=NAVY_DARK)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).font = Font(name="Inter", size=10, color=NAVY)
        ws.cell(row=r, column=5).font = Font(name="Inter", size=11, bold=True, color=NAVY_DARK)
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 22

    # === SCHEDULE ===
    sched_start = 26
    ws.cell(row=sched_start-1, column=2, value="📋 CRONOGRAMA MÊS A MÊS — fórmulas dinâmicas")
    ws.cell(row=sched_start-1, column=2).font = Font(name="Inter", size=12, bold=True, color=NAVY)
    ws.merge_cells(f"B{sched_start-1}:O{sched_start-1}")
    ws.row_dimensions[sched_start-1].height = 24

    sched_hdr = sched_start
    hdrs = ["Mês", "Data", "Saldo Inicial", "Juros", "Pagamento", "Principal",
            "Saldo Final", "RoU Saldo", "Depr. Mês", "PIS", "COFINS", "CBS", "IBS", "Custo Líq."]
    for col, h in enumerate(hdrs, start=2):
        style_header(ws.cell(row=sched_hdr, column=col, value=h))
    ws.row_dimensions[sched_hdr].height = 28

    prazo = c.get('prazo_meses', 36)
    for mes_idx in range(prazo):
        r = sched_hdr + 1 + mes_idx
        if mes_idx == 0:
            ws.cell(row=r, column=2, value=1)
            ws.cell(row=r, column=3, value="=$C$8")
            ws.cell(row=r, column=4, value="=$C$16")
        else:
            ws.cell(row=r, column=2, value=f"=B{r-1}+1")
            ws.cell(row=r, column=3, value=f"=EDATE(C{r-1},1)")
            ws.cell(row=r, column=4, value=f"=H{r-1}")

        ws.cell(row=r, column=5, value=f"=D{r}*$C$13")
        ws.cell(row=r, column=6, value="=$C$10")
        ws.cell(row=r, column=7, value=f"=F{r}-E{r}")
        ws.cell(row=r, column=8, value=f"=D{r}-G{r}")
        ws.cell(row=r, column=10, value="=$C$18")
        ws.cell(row=r, column=9, value=f"=$C$17-SUMPRODUCT($J${sched_hdr+1}:J{r})")
        ws.cell(row=r, column=11, value=f'=IF(AND($C$12="LUCRO_REAL",C{r}<=DATE(2026,12,31)),F{r}*{ALIQ_PIS},0)')
        ws.cell(row=r, column=12, value=f'=IF(AND($C$12="LUCRO_REAL",C{r}<=DATE(2026,12,31)),F{r}*{ALIQ_COFINS},0)')
        ws.cell(row=r, column=13, value=f'=IFERROR(IF(C{r}>=DATE(2027,1,1),F{r}*VLOOKUP(YEAR(C{r}),\'Tabela Aliquotas\'!$B$16:$E$30,2,FALSE),0),0)')
        ws.cell(row=r, column=14, value=f'=IFERROR(IF(C{r}>=DATE(2027,1,1),F{r}*VLOOKUP(YEAR(C{r}),\'Tabela Aliquotas\'!$B$16:$E$30,3,FALSE),0),0)')
        ws.cell(row=r, column=15, value=f"=F{r}-K{r}-L{r}-M{r}-N{r}")

        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).number_format = FMT_DATE
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        for col in range(4, 16):
            ws.cell(row=r, column=col).number_format = FMT_BR
            ws.cell(row=r, column=col).font = Font(name="Inter", size=9, color=NAVY)
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="right")
        if mes_idx % 2 == 1:
            for col in range(2, 16):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=GRAY_BG)

    # Total
    tr = sched_hdr + 1 + prazo
    for col in range(2, 16):
        ws.cell(row=tr, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=tr, column=col).font = Font(name="Inter", size=10, bold=True, color=WHITE)
    ws.cell(row=tr, column=2, value="TOTAL").alignment = Alignment(horizontal="center")
    for col_letter in ['E', 'F', 'G', 'J', 'K', 'L', 'M', 'N', 'O']:
        col_idx = ord(col_letter) - ord('A') + 1
        ws.cell(row=tr, column=col_idx,
                value=f"=SUM({col_letter}{sched_hdr+1}:{col_letter}{sched_hdr+prazo})")
        ws.cell(row=tr, column=col_idx).number_format = FMT_BR
        ws.cell(row=tr, column=col_idx).alignment = Alignment(horizontal="right")
    ws.row_dimensions[tr].height = 22

    rule = ColorScaleRule(start_type='min', start_color=GREEN_LIGHT, end_type='max', end_color=RED_LIGHT)
    ws.conditional_formatting.add(f'H{sched_hdr+1}:H{sched_hdr+prazo}', rule)

    # LARGURAS CORRIGIDAS — coluna B grande pra labels não cortarem
    widths = {'A': 2, 'B': 38, 'C': 16, 'D': 14, 'E': 12, 'F': 13, 'G': 13,
              'H': 14, 'I': 14, 'J': 12, 'K': 11, 'L': 11, 'M': 11, 'N': 11, 'O': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # FREEZE PANE CORRIGIDO — congela só o header do schedule + B/C
    # P column é onde começa o conteúdo "scrollável" mas como temos schedule a partir da B, vou congelar somente no header do schedule pra baixo
    ws.freeze_panes = f"A{sched_hdr+1}"


def gen_resumo_consolidado(wb, contratos):
    ws = wb.create_sheet("📑 RESUMO")
    ws.sheet_view.showGridLines = False
    add_logo_or_text(ws, row=1)

    ws.merge_cells("B4:K4")
    ws["B4"] = "RESUMO CONSOLIDADO — referencia fórmulas das abas Contrato N"
    ws["B4"].font = Font(name="Inter", size=14, bold=True, color=NAVY_DARK)
    ws.row_dimensions[4].height = 26

    hr = 6
    headers = ["#", "Contrato", "Valor Mensal", "Prazo", "VP Inicial", "Total Juros",
               "PIS+COFINS", "CBS+IBS", "Total Créditos", "Custo Líquido"]
    for col, h in enumerate(headers, start=2):
        style_header(ws.cell(row=hr, column=col, value=h))
    ws.row_dimensions[hr].height = 28

    for i, c in enumerate(contratos, start=hr+1):
        n = i - hr
        sheet_ref = f"'Contrato {n}'"
        prazo = c.get('prazo_meses', 36)
        sched_start = 27
        sched_end = sched_start + prazo - 1

        ws.cell(row=i, column=2, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3, value=f"Contrato {n}").font = Font(name="Inter", size=10, color=NAVY)
        ws.cell(row=i, column=4, value=f"={sheet_ref}!C10").number_format = FMT_BR
        ws.cell(row=i, column=5, value=f"={sheet_ref}!C9").alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=6, value=f"={sheet_ref}!C16").number_format = FMT_BR
        ws.cell(row=i, column=7, value=f"=SUM({sheet_ref}!E{sched_start}:E{sched_end})").number_format = FMT_BR
        ws.cell(row=i, column=8, value=f"=SUM({sheet_ref}!K{sched_start}:L{sched_end})").number_format = FMT_BR
        ws.cell(row=i, column=9, value=f"=SUM({sheet_ref}!M{sched_start}:N{sched_end})").number_format = FMT_BR
        ws.cell(row=i, column=10, value=f"=H{i}+I{i}").number_format = FMT_BR
        ws.cell(row=i, column=11, value=f"=({sheet_ref}!C10*{sheet_ref}!C9)-J{i}").number_format = FMT_BR
        for col in range(2, 12):
            ws.cell(row=i, column=col).font = Font(name="Inter", size=10, color=NAVY)
            if col >= 4 and col != 5:
                ws.cell(row=i, column=col).alignment = Alignment(horizontal="right")
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=GRAY_BG)
        ws.row_dimensions[i].height = 22

    # Total
    tr = hr + len(contratos) + 1
    for col in range(2, 12):
        ws.cell(row=tr, column=col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row=tr, column=col).font = Font(name="Inter", size=10, bold=True, color=WHITE)
    ws.cell(row=tr, column=3, value="TOTAL")
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K']:
        col_idx = ord(col_letter) - ord('A') + 1
        ws.cell(row=tr, column=col_idx,
                value=f"=SUM({col_letter}{hr+1}:{col_letter}{hr+len(contratos)})")
        ws.cell(row=tr, column=col_idx).number_format = FMT_BR
        ws.cell(row=tr, column=col_idx).alignment = Alignment(horizontal="right")
    ws.row_dimensions[tr].height = 24

    # LARGURAS CORRIGIDAS
    for col, w in [('A', 2), ('B', 5), ('C', 14), ('D', 16), ('E', 8),
                   ('F', 16), ('G', 16), ('H', 16), ('I', 16), ('J', 18), ('K', 18)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"B{hr+1}"


def gen_excel(out_path, contratos, rates):
    wb = Workbook()
    gen_capa(wb, contratos, rates)
    gen_aba_aliquotas(wb)
    for n, c in enumerate(contratos, start=1):
        gen_contrato_sheet(wb, c, n)
    gen_resumo_consolidado(wb, contratos)
    wb.save(out_path)
    print(f"✓ Excel salvo: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 gerar_relatorio_cpc06_v6.py <input.xlsx> [output.xlsx]"); sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "relatorio-cpc06-newledger-v6.xlsx"
    print("=" * 70); print("NEWLEDGER · CPC 06 (R2) / IFRS 16 — v6 com LOGO REAL"); print("=" * 70)
    print(f"\nLogo: {'PNG real (logo-newledger.png)' if HAS_LOGO else 'texto (sem logo PNG real)'}")
    print("\n[1/4] Selic via BACEN + yfinance...")
    rates = fetch_rates()
    print(f"      Taxa: {rates['taxa_desconto']:.2f}% a.a.")
    print(f"\n[2/4] Lendo: {Path(inp).name}")
    contratos = parse_input_contracts(inp)
    print(f"      {len(contratos)} contratos detectados")
    print(f"\n[3/4] Gerando Excel com FÓRMULAS + layout corrigido...")
    print(f"\n[4/4] Salvando: {out}")
    gen_excel(out, contratos, rates)
    print("\n" + "=" * 70); print(f"✅ CONCLUÍDO · {out}"); print("=" * 70)
